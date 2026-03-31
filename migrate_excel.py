"""
Migrate data from Easy Life - Management Board.xlsx to Supabase.
Run with: C:/Users/Chris-Optinet/AppData/Local/Programs/Python/Python312/python.exe migrate_excel.py
"""

import sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
import openpyxl
from supabase import create_client
from datetime import datetime

SUPABASE_URL = "https://vdhoiqyalblwrnwbprcr.supabase.co"
SUPABASE_KEY = "sb_publishable_X_NfeKkcQcmyxOGSLSMOCg_vVOcox0W"

client = create_client(SUPABASE_URL, SUPABASE_KEY)
wb = openpyxl.load_workbook("Easy Life - Management Board.xlsx", data_only=True)

def fmt_date(v):
    if v is None: return None
    if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
    if isinstance(v, str):
        v = v.strip()
        if not v or v == '-': return None
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
            try: return datetime.strptime(v, fmt).strftime("%Y-%m-%d")
            except: pass
    return None

def clean(v):
    if v is None: return None
    if isinstance(v, str) and v.strip() in ('-', '', 'None'): return None
    return v

def num(v):
    if v is None: return None
    try: return float(v)
    except: return None

def row_vals(ws, row_idx, max_col=None):
    row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True))[0]
    return row if not max_col else row[:max_col]

print("\n=== MIGRATING GUESTS ===")
ws = wb['Guests']
guests = []
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[2]: continue  # no check_in
    room = row[1]
    if not room: continue
    try:
        room_int = int(room)
    except:
        continue
    check_in = fmt_date(row[2])
    check_out = fmt_date(row[3])
    if not check_in or not check_out: continue
    nights_raw = row[4]  # formula result or None
    nights = None
    try:
        if isinstance(nights_raw, (int, float)):
            nights = int(nights_raw)
        elif check_in and check_out:
            from datetime import date
            d1 = datetime.strptime(check_in, "%Y-%m-%d").date()
            d2 = datetime.strptime(check_out, "%Y-%m-%d").date()
            nights = (d2 - d1).days
    except: pass

    amount_day = num(row[7])
    amount_stay = num(row[8])
    if amount_day and nights and not amount_stay:
        amount_stay = amount_day * nights
    payment = num(row[10]) or 0.0

    g = {
        "date": fmt_date(row[0]),
        "room": room_int,
        "check_in": check_in,
        "check_out": check_out,
        "guest_name": str(row[5]).strip() if row[5] else "Unknown",
        "guest_count": int(row[6]) if row[6] and str(row[6]).strip() != '-' else 1,
        "amount_thb_day": amount_day,
        "amount_thb_stay": amount_stay,
        "paid": str(clean(row[9])) if clean(row[9]) else None,
        "payment": payment,
        "invoice": str(clean(row[12])) if clean(row[12]) else None,
        "notes": str(clean(row[13])) if clean(row[13]) else None,
        "email": str(clean(row[14])) if clean(row[14]) else None,
        "phone": str(clean(row[15])) if clean(row[15]) else None,
        "tm30": str(row[16]).lower() in ('yes', 'true', '1') if row[16] else False,
    }
    guests.append(g)

print(f"  Found {len(guests)} guests")
if guests:
    res = client.table("guests").insert(guests).execute()
    print(f"  Inserted {len(res.data)} guests OK")


print("\n=== MIGRATING EXPENSES ===")
ws = wb['Expenses']
expenses = []
for row in ws.iter_rows(min_row=5, values_only=True):
    if not row[5]: continue  # no transaction number
    amount = num(row[10])
    if amount is None: continue
    e = {
        "audit": str(clean(row[0])) if clean(row[0]) else None,
        "lawyers": str(clean(row[1])) if clean(row[1]) else None,
        "sent": fmt_date(row[2]),
        "to_verify": str(clean(row[3])) if clean(row[3]) else None,
        "payment_date": fmt_date(row[4]),
        "transaction_number": str(row[5]).strip() if row[5] else None,
        "document_number": str(clean(row[6])) if clean(row[6]) else None,
        "category": str(clean(row[7])) if clean(row[7]) else None,
        "subcategory": str(clean(row[8])) if clean(row[8]) else None,
        "supplier": str(clean(row[9])) if clean(row[9]) else None,
        "amount": amount,
        "currency": str(row[11]).strip() if row[11] else "THB",
        "method": str(clean(row[12])) if clean(row[12]) else None,
        "paid_by": str(clean(row[13])) if clean(row[13]) else None,
        "internal_document": str(clean(row[14])) if clean(row[14]) else None,
        "document_page": str(clean(row[15])) if clean(row[15]) else None,
        "type": str(clean(row[16])) if clean(row[16]) else None,
        "description": str(clean(row[17])) if clean(row[17]) else None,
        "is_legal": False,
    }
    expenses.append(e)

print(f"  Found {len(expenses)} expenses")
if expenses:
    # Insert in batches of 50
    for i in range(0, len(expenses), 50):
        batch = expenses[i:i+50]
        res = client.table("expenses").insert(batch).execute()
    print(f"  Inserted {len(expenses)} expenses OK")


print("\n=== MIGRATING EXPENSES (LEGAL) ===")
ws = wb['Expenses Shared with Legal']
legal_expenses = []
existing_refs = {e['transaction_number'] for e in expenses if e['transaction_number']}

for row in ws.iter_rows(min_row=6, values_only=True):
    if not row[4]: continue  # no transaction number
    txn = str(row[4]).strip()
    amount = num(row[8]) if len(row) > 8 else None
    if amount is None: continue
    # Skip if already in expenses (duplicate)
    if txn in existing_refs: continue
    e = {
        "audit": None,
        "lawyers": str(clean(row[0])) if clean(row[0]) else None,
        "sent": fmt_date(row[1]),
        "to_verify": str(clean(row[2])) if clean(row[2]) else None,
        "payment_date": fmt_date(row[3]),
        "transaction_number": txn,
        "document_number": str(clean(row[5])) if clean(row[5]) else None,
        "category": str(clean(row[6])) if clean(row[6]) else None,
        "subcategory": str(clean(row[7])) if clean(row[7]) else None,
        "supplier": str(row[8]).strip() if row[8] else None,
        "amount": num(row[9]) if len(row) > 9 else amount,
        "currency": str(row[10]).strip() if len(row) > 10 and row[10] else "THB",
        "method": str(clean(row[11])) if len(row) > 11 and clean(row[11]) else None,
        "paid_by": str(clean(row[12])) if len(row) > 12 and clean(row[12]) else None,
        "internal_document": None,
        "document_page": None,
        "type": None,
        "description": str(clean(row[13])) if len(row) > 13 and clean(row[13]) else None,
        "is_legal": True,
    }
    legal_expenses.append(e)

print(f"  Found {len(legal_expenses)} legal-only expenses")
if legal_expenses:
    for i in range(0, len(legal_expenses), 50):
        batch = legal_expenses[i:i+50]
        res = client.table("expenses").insert(batch).execute()
    print(f"  Inserted {len(legal_expenses)} legal expenses OK")


print("\n=== MIGRATING REVENUE ===")
ws = wb['Revenue']
revenues = []
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[0]: continue
    d = fmt_date(row[0])
    if not d: continue
    amount = num(row[3])
    if amount is None: continue
    revenues.append({
        "date": d,
        "type": str(clean(row[1])) if clean(row[1]) else None,
        "supplier": str(clean(row[2])) if clean(row[2]) else None,
        "amount_thb": amount,
        "notes": None,
    })

print(f"  Found {len(revenues)} revenue records")
if revenues:
    res = client.table("revenue").insert(revenues).execute()
    print(f"  Inserted {len(res.data)} revenue records OK")


print("\n=== MIGRATING FOUNDING CONTRIBUTIONS ===")
ws = wb['Founding']
SHAREHOLDER_COLS = {
    2: "Lorenzo PAGNAN",
    3: "Stella MAROZZI",
    4: "Bruce MIFSUD",
    5: "Hanna PARSONSON",
}
contributions = []
for row in ws.iter_rows(min_row=5, values_only=True):
    if not row[0]: continue
    d = fmt_date(row[0])
    if not d: continue
    method = str(clean(row[1])) if clean(row[1]) else None
    notes = str(clean(row[6])) if len(row) > 6 and clean(row[6]) else None
    for col_idx, shareholder in SHAREHOLDER_COLS.items():
        amount = num(row[col_idx])
        if amount is None or amount == 0: continue
        try:
            float_val = float(str(amount).strip())
            if float_val == 0: continue
        except: continue
        contributions.append({
            "date": d,
            "method": method,
            "shareholder": shareholder,
            "amount_thb": abs(float(str(amount))),
            "amount_eur": None,
            "notes": notes,
        })

print(f"  Found {len(contributions)} founding contributions")
if contributions:
    for i in range(0, len(contributions), 50):
        res = client.table("founding_contributions").insert(contributions[i:i+50]).execute()
    print(f"  Inserted {len(contributions)} contributions OK")


print("\n=== MIGRATING SHAREHOLDER WORK ===")
ws = wb['Shareholder Work']
WORK_COLS = {
    2: "Lorenzo PAGNAN",
    4: "Stella MAROZZI",
    6: "Bruce MIFSUD",
    8: "Hanna PARSONSON",
}
work_rows = []
for row in ws.iter_rows(min_row=6, values_only=True):
    if not row[0]: continue
    d = fmt_date(row[0])
    if not d: continue
    rate = num(row[1]) or 200
    for col_idx, shareholder in WORK_COLS.items():
        hours = num(row[col_idx])
        if hours is None or hours == 0: continue
        work_rows.append({
            "month": d,
            "shareholder": shareholder,
            "hours": float(hours),
            "hour_rate": float(rate),
        })

print(f"  Found {len(work_rows)} shareholder work records")
if work_rows:
    res = client.table("shareholder_work").insert(work_rows).execute()
    print(f"  Inserted {len(res.data)} work records OK")


print("\n=== MIGRATING TASKS (TO DO) ===")
ws = wb['To Do']
todos = []
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[2]: continue  # no topic
    topic = str(row[2]).strip()
    if not topic or topic == 'TOPIC': continue
    status = str(row[6]).strip() if row[6] else 'Pending'
    todos.append({
        "project": str(clean(row[0])) if clean(row[0]) else "EasyLife",
        "department": str(clean(row[1])) if clean(row[1]) else None,
        "topic": topic,
        "responsible_person": str(clean(row[3])) if clean(row[3]) else None,
        "status_notes": str(clean(row[4])) if clean(row[4]) else None,
        "target_date": fmt_date(row[5]),
        "status": status if status in ('Pending', 'Ongoing', 'Complete', 'Blocked') else 'Pending',
    })

print(f"  Found {len(todos)} tasks")
if todos:
    res = client.table("todos").insert(todos).execute()
    print(f"  Inserted {len(res.data)} tasks OK")


print("\n=== MIGRATING BUDGET RENT ===")
ws = wb['Budget - Rent']
rent_rows = []
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[0]: continue
    year_num = num(row[0])
    if year_num is None: continue
    rent = num(row[2])
    if rent is None: continue
    vat = float(rent) * 0.0266
    rent_rows.append({
        "year_number": int(year_num),
        "year_label": str(clean(row[1])) if clean(row[1]) else f"Rent {2024 + int(year_num)}",
        "rent_thb": float(rent),
        "vat_amount": round(vat, 2),
    })

print(f"  Found {len(rent_rows)} rent schedule rows")
if rent_rows:
    res = client.table("budget_rent").insert(rent_rows).execute()
    print(f"  Inserted {len(res.data)} rent rows OK")


print("\n=== MIGRATING BUDGET REVENUE ===")
ws = wb['Budget - Revenue']
budget_rev = []
MONTH_MAP = {1:'Jan',2:'Feb',3:'Mar',4:'Apr',5:'May',6:'Jun',7:'Jul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
# Row 3 has year, row 4+ has room data, columns B-M = months 1-12
year_row = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
budget_year = int(year_row[0]) if year_row[0] and str(year_row[0]).isdigit() else 2026

for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[0]: continue
    room_name = str(row[0]).strip()
    if not room_name or room_name in ('Room', ''): continue
    for mi in range(12):
        amount = num(row[mi + 1])
        if amount and amount > 0:
            budget_rev.append({
                "year": budget_year,
                "month": mi + 1,
                "room_name": room_name,
                "amount_thb": float(amount),
                "season": None,
            })

print(f"  Found {len(budget_rev)} budget revenue entries")
if budget_rev:
    for i in range(0, len(budget_rev), 100):
        client.table("budget_revenue").insert(budget_rev[i:i+100]).execute()
    print(f"  Inserted {len(budget_rev)} budget revenue entries OK")


print("\n=== MIGRATING BUDGET EXPENSES ===")
ws = wb['Budget - Expenses']
budget_exp = []
year_row2 = list(ws.iter_rows(min_row=3, max_row=3, values_only=True))[0]
budget_year2 = int(year_row2[0]) if year_row2[0] and str(year_row2[0]).strip().isdigit() else 2026

current_category = ''
current_type = 'OPEX'
for row in ws.iter_rows(min_row=4, values_only=True):
    if not row[0]: continue
    item_name = str(row[0]).strip()
    if not item_name: continue
    # Detect category headers (all caps or known patterns)
    if item_name in ('OPERATING COSTS (OPEX)', 'CAPITAL EXPENDITURE (CAPEX)'):
        current_type = 'CAPEX' if 'CAPEX' in item_name else 'OPEX'
        continue
    # Detect subcategory rows (no amounts, just label)
    has_amounts = any(num(row[mi + 1]) is not None for mi in range(12) if mi + 1 < len(row))
    if not has_amounts: continue
    for mi in range(12):
        if mi + 1 >= len(row): break
        amount = num(row[mi + 1])
        if amount and amount > 0:
            budget_exp.append({
                "year": budget_year2,
                "month": mi + 1,
                "category": current_type,
                "item_name": item_name,
                "amount_thb": float(amount),
                "expense_type": current_type,
            })

print(f"  Found {len(budget_exp)} budget expense entries")
if budget_exp:
    for i in range(0, len(budget_exp), 100):
        client.table("budget_expenses").insert(budget_exp[i:i+100]).execute()
    print(f"  Inserted {len(budget_exp)} budget expense entries OK")


print("\n=== MIGRATING BANK BALANCES (Status of Account) ===")
ws = wb['Status_of_Account']
bank_rows = []
# Rows 8-10 have Bank, Nicolo, No Invoice/Receipt with amounts
for row in ws.iter_rows(min_row=8, max_row=10, values_only=True):
    if not row[0]: continue
    label = str(row[0]).strip()
    amount = num(row[1])
    if amount is None: continue
    d = fmt_date(row[4]) if len(row) > 4 else None
    status = str(row[5]).strip() if len(row) > 5 and row[5] else None
    bank_rows.append({
        "label": label,
        "amount": float(amount),
        "recorded_date": d or "2026-03-30",
        "status": status,
        "notes": None,
    })

print(f"  Found {len(bank_rows)} bank balance records")
if bank_rows:
    res = client.table("bank_balances").insert(bank_rows).execute()
    print(f"  Inserted {len(res.data)} bank balance records OK")


print("\nDONE Migration complete!")
