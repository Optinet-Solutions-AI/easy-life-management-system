import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import re
import json
from collections import defaultdict

FILE_PATH = r"c:\Users\Chris-Optinet\elms\Easy Life - Management Board.xlsx"

# Load workbook - keep formulas as strings
wb = load_workbook(FILE_PATH, data_only=False)
wb_data = load_workbook(FILE_PATH, data_only=True)

print("=" * 80)
print("SHEET NAMES")
print("=" * 80)
for name in wb.sheetnames:
    print(f"  - {name}")

print()

# Regex to find cross-sheet references like 'SheetName'!A1 or SheetName!A1
cross_ref_pattern = re.compile(r"'?([^'!]+)'?![\$A-Z][\$0-9A-Z:]+", re.IGNORECASE)

all_dependencies = defaultdict(set)  # sheet -> set of sheets it references

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws_data = wb_data[sheet_name]

    print("=" * 80)
    print(f"SHEET: {sheet_name}")
    print("=" * 80)

    # Dimensions
    print(f"  Dimensions: {ws.dimensions}")
    print(f"  Max Row: {ws.max_row}, Max Col: {ws.max_column}")

    # Collect headers (first row)
    print("\n  --- ROW 1 (Headers/First Row) ---")
    first_row_vals = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        val = cell.value
        if val is not None:
            first_row_vals.append(f"    Col {get_column_letter(col)} ({col}): {repr(val)}")
    if first_row_vals:
        for v in first_row_vals:
            print(v)
    else:
        print("    (empty)")

    # Collect all non-empty cell data, formulas, and references
    print("\n  --- ALL NON-EMPTY CELLS (first 200) ---")
    cell_count = 0
    formulas_found = []
    cross_sheet_refs = defaultdict(list)  # ref_sheet -> list of (cell_addr, formula)
    data_types = defaultdict(int)  # type -> count

    sample_data = []  # first 5 rows of actual data

    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell_count += 1
                val_str = str(cell.value)

                # Track data type
                if val_str.startswith("="):
                    data_types["formula"] += 1
                    formulas_found.append((cell.coordinate, val_str))
                    # Find cross-sheet refs
                    refs = cross_ref_pattern.findall(val_str)
                    for ref in refs:
                        cross_sheet_refs[ref].append((cell.coordinate, val_str))
                        all_dependencies[sheet_name].add(ref)
                elif isinstance(cell.value, (int, float)):
                    data_types["number"] += 1
                elif hasattr(cell.value, 'strftime'):
                    data_types["date"] += 1
                else:
                    data_types["text"] += 1

                if cell_count <= 200:
                    data_val = ws_data[cell.coordinate].value if ws_data[cell.coordinate] else None
                    sample_data.append({
                        "cell": cell.coordinate,
                        "formula": val_str if val_str.startswith("=") else None,
                        "value": val_str if not val_str.startswith("=") else None,
                        "computed": data_val
                    })

    print(f"  Total non-empty cells: {cell_count}")
    print(f"  Data types: {dict(data_types)}")

    # Print sample rows (rows 1-10)
    print("\n  --- SAMPLE DATA (rows 1-10, all columns) ---")
    for row_idx in range(1, min(11, ws.max_row + 1)):
        row_data = []
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            data_cell = ws_data.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                row_data.append(f"{cell.coordinate}={repr(cell.value)} [computed={repr(data_cell.value)}]")
        if row_data:
            print(f"  Row {row_idx}: {' | '.join(row_data)}")

    # Print ALL unique formulas
    if formulas_found:
        print(f"\n  --- FORMULAS ({len(formulas_found)} total) ---")
        seen_formula_types = {}
        for coord, formula in formulas_found:
            # Normalize to show unique patterns
            norm = re.sub(r'[A-Z]+\d+', 'REF', formula)
            if norm not in seen_formula_types:
                seen_formula_types[norm] = (coord, formula)

        for norm, (coord, formula) in list(seen_formula_types.items())[:50]:
            print(f"    {coord}: {formula}")

    # Cross-sheet references
    if cross_sheet_refs:
        print(f"\n  --- CROSS-SHEET REFERENCES ---")
        for ref_sheet, refs in cross_sheet_refs.items():
            print(f"  References sheet '{ref_sheet}':")
            for coord, formula in refs[:5]:  # show up to 5 examples
                print(f"    {coord}: {formula}")
            if len(refs) > 5:
                print(f"    ... and {len(refs)-5} more")

    print()

print("=" * 80)
print("CROSS-SHEET DEPENDENCY MAP")
print("=" * 80)
for sheet, deps in all_dependencies.items():
    print(f"  '{sheet}' reads from: {deps}")

print()
print("=" * 80)
print("NAMED RANGES / DEFINED NAMES")
print("=" * 80)
if wb.defined_names:
    for name, defn in wb.defined_names.items():
        print(f"  {name}: {defn.attr_text if hasattr(defn, 'attr_text') else defn}")
else:
    print("  None found")
